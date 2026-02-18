from PySide6.QtWidgets import QLabel, QVBoxLayout, QComboBox, QMessageBox, QSizePolicy
from openpyxl import load_workbook
from gui.steps.step_widget import StepWidget
from gui.widgets.file_picker import FilePickerWidget


class StepFileSelect(StepWidget):
    def setup_ui(self):
        self.title = "Кезең 1 / 5: Бастапқы деректерді таңдау"
        self.description = (
            "Файлды таңдаңыз, өңделетін парақты (лист) және топты көрсетіңіз."
        )

        self.layout = QVBoxLayout(self)

        # Excel файлды таңдау
        self.file_select_widget = FilePickerWidget("Мониторинг файлы")
        self.file_select_widget.setSizePolicy(
            QSizePolicy.Expanding, QSizePolicy.Preferred
        )

        # Sheet таңдау
        self.combo_sheet = QComboBox()
        self.combo_sheet.setEnabled(False)

        sheet_label = QLabel("Кесте (sheet) атауы:")
        sheet_label.setBuddy(self.combo_sheet)
        sheet_label.setProperty("lbl-level", "lbl")

        # Топ түрі

        self.combo_group = QComboBox()
        self.group_types = {
            "early_age": "Бөбек (ерте жас)",
            "junior": "Кіші топ",
            "middle": "Ортаңғы топ",
            "senior": "Ересек топ",
            "preschool": "Мектепке даярлық тобы",
        }
        for key, display_text in self.group_types.items():
            self.combo_group.addItem(display_text, key)

        group_type_label = QLabel("Топ түрі:")
        group_type_label.setBuddy(self.combo_group)
        group_type_label.setProperty("lbl-level", "lbl")

        # ---- Layout ----
        self.layout.addWidget(self.file_select_widget)
        self.layout.addSpacing(10)

        self.layout.addWidget(sheet_label)
        self.layout.addWidget(self.combo_sheet)
        self.layout.addSpacing(10)

        self.layout.addWidget(group_type_label)
        self.layout.addWidget(self.combo_group)
        self.layout.addStretch()

    def setup_state_machine(self):
        return

    def connect_signals(self):
        self.file_select_widget.fileSelected.connect(self.select_excel_file)

    def select_excel_file(self, file_path):
        try:
            wb = load_workbook(file_path, read_only=True)
            self.state.workbook = wb
            self.combo_sheet.clear()
            self.combo_sheet.addItems(wb.sheetnames)
            self.combo_sheet.setEnabled(True)
        except Exception as e:
            QMessageBox.critical(self, "Қате", f"Файлды оқу кезінде қате:\n{e}")
            self.combo_sheet.setEnabled(False)

    def validate_before_next(self):
        if not self.state.workbook:
            QMessageBox.warning(self, "Ескерту", "Excel файлын таңдаңыз.")
            return False

        if not self.combo_sheet.currentText():
            QMessageBox.warning(self, "Ескерту", "Кесте (sheet) атауын таңдаңыз.")
            return False

        self.state.sheet_name = self.combo_sheet.currentText()
        self.state.age_group = self.combo_group.currentData()

        return True
