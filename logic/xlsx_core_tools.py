import re
from openpyxl.utils.cell import range_boundaries


def _fix_merged_cells_before_deletion(sheet, row_idx):
    """
    Жол жойылмас бұрын біріккен ұяшықтарды ЖОҒАРЫДАН ТӨМЕН ҚАРАЙ сорттап,
    ішкі XML елестерін қатесіз, ретімен тазалайды және қайта құрады.
    """

    # МАҢЫЗДЫ: Жолдың өсу ретімен міндетті түрде сұрыптау
    merged_ranges = sorted(
        list(sheet.merged_cells.ranges), key=lambda r: (r.min_row, r.min_col)
    )

    # print("-----")
    for m_range in merged_ranges:
        coord = m_range.coord
        min_col, min_row, max_col, max_row = range_boundaries(coord)
        # print(coord, min_col, min_row, max_col, max_row)

        # Шарт А: Егер жойылатын жол біріккен ұяшықтың ИНТЕРВАЛЫНА КІРСЕ
        if min_row <= row_idx <= max_row:
            try:
                sheet.unmerge_cells(coord)
            except ValueError:
                pass

            if m_range in sheet.merged_cells.ranges:
                sheet.merged_cells.ranges.remove(m_range)

            new_min_row = None
            new_max_row = None

            # 1-жағдай: Вертикальді үлкен аймақ
            if max_row > min_row:
                if row_idx == max_row:
                    new_max_row = max_row - 1
                    new_min_row = min_row
                elif row_idx == min_row:
                    new_min_row = min_row + 1
                    new_max_row = max_row
                else:
                    new_max_row = max_row - 1
                    new_min_row = min_row

            # 2-жағдай: Горизонтальді 1 жолдық аймақ
            else:
                if row_idx != min_row:
                    new_min_row = min_row
                    new_max_row = max_row

            # Қайта біріктіру шарттары
            if new_min_row is not None and new_max_row is not None:
                if new_max_row >= new_min_row and (
                    new_max_row > new_min_row or max_col > min_col
                ):
                    sheet.merge_cells(
                        start_row=new_min_row,
                        start_column=min_col,
                        end_row=new_max_row,
                        end_column=max_col,
                    )

        # Шарт Ә: Егер біріккен ұяшық жойылатын жолдан ТӨМЕН тұрса
        elif row_idx < min_row:
            try:
                sheet.unmerge_cells(coord)
            except ValueError:
                pass
            if m_range in sheet.merged_cells.ranges:
                sheet.merged_cells.ranges.remove(m_range)

            # Төмендегі ұяшықты 1 жолға қауіпсіз жоғары көшіру
            sheet.merge_cells(
                start_row=min_row - 1,
                start_column=min_col,
                end_row=max_row - 1,
                end_column=max_col,
            )


def _shift_formulas_after_deletion(sheet, row_idx):
    """Жол жойылғаннан кейін формула сілтемелерін таза математикалық жолмен 1-ге азайтады."""
    cell_pattern = re.compile(r"([A-Z]+)(\d+)")

    for row in range(row_idx, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)

            if cell.value and str(cell.value).startswith("="):
                formula = str(cell.value)

                def update_formula_refs(match):
                    col_letter = match.group(1)
                    row_num = int(match.group(2))

                    # Егер формула сілтеме жасаған жол жойылған жолдан ТӨМЕН болса, жай ғана 1 жол жоғары жылжытамыз
                    if row_num > row_idx:
                        return f"{col_letter}{row_num - 1}"
                    return f"{col_letter}{row_num}"

                new_formula = cell_pattern.sub(update_formula_refs, formula)
                if new_formula != formula:
                    cell.value = new_formula


def _shift_row_heights_after_deletion(sheet, row_idx, default_height=15.75):
    max_row = sheet.max_row

    for r in range(row_idx, max_row + 1):
        next_row = r + 1

        # Келесі жолдың өлшемін көшіреміз
        if next_row in sheet.row_dimensions:
            sheet.row_dimensions[r].height = sheet.row_dimensions[next_row].height
            # МАҢЫЗДЫ: Жасырын күйін де тасымалдаймыз
            sheet.row_dimensions[r].hidden = sheet.row_dimensions[next_row].hidden
        else:
            sheet.row_dimensions[r].height = None
            sheet.row_dimensions[r].hidden = False

        # САҚТАНДЫРҒЫШ: Егер жылжығаннан кейін қазіргі көрінуі тиіс жолдың
        # биіктігі тым кішкентай (мысалы, 5-тен кіші) немесе жасырын болып қалса, оны түзетеміз
        if r >= row_idx and (
            sheet.row_dimensions[r].height is not None
            and sheet.row_dimensions[r].height < 5
        ):
            sheet.row_dimensions[r].height = default_height
            sheet.row_dimensions[r].hidden = False

    # Ең соңғы жолды реттеу
    sheet.row_dimensions[max_row + 1].height = default_height
    sheet.row_dimensions[max_row + 1].hidden = False


def delete_row(sheet, row_idx):
    if row_idx < 1 or row_idx > sheet.max_row:
        return

    # 1. Біріккен ұяшықтарды реттеу
    _fix_merged_cells_before_deletion(sheet, row_idx)

    # 2. Жолды физикалық түрде өшіру
    sheet.delete_rows(row_idx, amount=1)

    # 3. Жол биіктіктерін Excel сияқты қалпына келтіру
    _shift_row_heights_after_deletion(sheet, row_idx, default_height=15.75)

    # 4. Формулаларды қауіпсіз жылжыту
    _shift_formulas_after_deletion(sheet, row_idx)
