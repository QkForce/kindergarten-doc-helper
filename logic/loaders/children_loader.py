from openpyxl.worksheet.worksheet import Worksheet
from typing import Tuple, Optional, List
from logic.loaders.utils import normalize_text, fuzzy_ratio, is_empty


class ChildrenLoader:
    def __init__(
        self,
        sheet,
        header_candidates: Optional[List[str]] = None,
        threshold: float = 0.78,
    ):
        self.sheet = sheet
        # header candidates in kazakh/russian variations you expect
        self.header_candidates = header_candidates or [
            "баланың аты жөні",
            "балалардың аты-жөні",
            "балалардың аты жөні",
            "аты-жөні",
            "аты жөні",
            "тегі және аты",
            "ф.и.о",
            "фамилия имя отчество",
            "балалар",
        ]
        self.stop_words = [
            "барлығы",
            "қорытынды",
            "ескерту",
            "жоғары",
            "орташа",
            "төмен",
        ]
        self.header_candidates = [normalize_text(x) for x in self.header_candidates]
        self.threshold = threshold

    def is_stop_word(self, value: any) -> bool:
        txt = normalize_text(value)
        if not txt:
            return False
        # Кез келген тоқтату сөзі ұяшық мәтінінің ішінде кездессе
        return any(word in txt for word in self.stop_words)

    def detect_header(self, sheet: Worksheet) -> Optional[Tuple[int, int]]:
        for row in sheet.iter_rows(
            min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column
        ):
            for cell in row:
                txt = normalize_text(cell.value)
                if not txt:
                    continue
                for cand in self.header_candidates:
                    if fuzzy_ratio(txt, cand) >= self.threshold:
                        return cell.row, cell.column
        return None, None

    def detect(self, sheet: Worksheet) -> Optional[Tuple[int, int, int]]:
        header_row, header_col = self.detect_header(sheet)

        # 1. start_row табу
        for r in range(header_row + 1, sheet.max_row):
            if not is_empty(sheet, r, header_col) and not self.is_stop_word(
                sheet.cell(row=r, column=header_col).value
            ):
                start_row = r
                break
        else:
            raise RuntimeError("The row where the list starts was not found")

        # 2. end_row табу
        for r in range(start_row + 1, sheet.max_row):
            if is_empty(sheet, r, header_col) or self.is_stop_word(
                sheet.cell(row=r, column=header_col).value
            ):
                end_row = r - 1
                break
        else:
            end_row = sheet.max_row

        # 3. name_col
        name_col = header_col
        return start_row, end_row, name_col

    def load_manual(self, start_row: int, end_row: int, name_col: int):
        children = []
        for row in self.sheet.iter_rows(
            min_row=start_row, max_row=end_row, values_only=True
        ):
            name = row[name_col - 1]
            children.append(name or "")
        return (
            children,
            start_row,
            end_row,
            name_col,
        )

    def load_auto(self):
        start_row, end_row, name_col = self.detect(self.sheet)
        return self.load_manual(start_row, end_row, name_col)
