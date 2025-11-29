from openpyxl.worksheet.worksheet import Worksheet
from logic.loaders.utils import (
    is_metric_code,
    find_next_nonempty_row,
    normalize_code_text,
)


class MetricsLoader:
    COL_STEP = 3

    def __init__(self, sheet: Worksheet):
        self.sheet = sheet

    def detect(self, sheet: Worksheet, max_col):
        metric_code_row = None
        metric_code_col = None

        # 1) Метрика кодын табу
        for row in range(1, sheet.max_row + 1):
            for col in range(1, max_col + 1):
                val = sheet.cell(row=row, column=col).value
                if is_metric_code(val):
                    metric_code_row = row
                    metric_code_col = col
                    break
            if metric_code_row:
                break

        if metric_code_row is None:
            raise ValueError("Метрика коды табылмады!")

        # 2) Метрика сипаттамасының жолын табу
        metric_desc_row = find_next_nonempty_row(
            sheet, metric_code_row, metric_code_col
        )
        if metric_desc_row is None:
            raise ValueError("Метрика сипаттамасы табылмады!")

        # 3) Горизонталь диапазон соңын табу
        metrics_start_col = metric_code_col
        metrics_end_col = self._calc_end_col(sheet, metric_code_row, metric_code_col)

        return (
            metric_code_row,
            metric_desc_row,
            metrics_start_col,
            metrics_end_col,
        )

    def _calc_end_col(self, sheet, metric_code_row, start_col):
        col = start_col
        last_good_col = start_col

        while col <= sheet.max_column:
            cell_val = sheet.cell(row=metric_code_row, column=col).value
            txt = normalize_code_text(cell_val)

            if txt == "" or not is_metric_code(txt):
                break

            last_good_col = col
            col += self.COL_STEP

        return last_good_col

    def load_manual(self, code_row: int, desc_row: int, start_col: int, end_col: int):
        metrics = []
        for col in range(start_col, end_col + 2, self.COL_STEP):
            code = self.sheet.cell(row=code_row, column=col).value
            desc = self.sheet.cell(row=desc_row, column=col).value
            metrics.append({"code": normalize_code_text(code), "desc": desc or ""})
        return metrics, code_row, desc_row, start_col, end_col

    def load_auto(self):
        max_col = int(self.sheet.max_column / 30)
        code_row, desc_row, start_col, end_col = self.detect(self.sheet, max_col)
        return self.load_manual(code_row, desc_row, start_col, end_col)
