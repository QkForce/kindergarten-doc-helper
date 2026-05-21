import re
from time import sleep
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from config.config import AGE_GROUP_DATA
from logic.xlsx_core_tools import delete_row
from logic.xlsx_detectors import (
    find_table_origin,
    find_student_list_start_row,
    find_student_name_col_index,
    find_footer_row_index,
    find_last_data_col_index,
    detect_actual_student_count,
)

METRIC_CODES = [
    code
    for metric_group_data in AGE_GROUP_DATA.values()
    for code in metric_group_data.keys()
]


def get_sheet_names(file_path: str):
    workbook = load_workbook(file_path, read_only=True)
    try:
        return workbook.sheetnames
    finally:
        workbook.close()


def load_metrics_from_excel(file_path, sheet_name, min_row=14, max_row=38):
    workbook = load_workbook(filename=file_path, read_only=False)
    sheet = workbook[sheet_name]

    children_data = []
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, values_only=True):
        child = {
            "number": row[0],
            "name": row[1],
        }
        for index, metrics_key in enumerate(METRIC_CODES):
            if row[2 + index * 3] is not None:
                child[metrics_key] = 1
            elif row[2 + index * 3 + 1] is not None:
                child[metrics_key] = 2
            elif row[2 + index * 3 + 2] is not None:
                child[metrics_key] = 3
        children_data.append(child)
    return children_data


def fill_assessment_table(
    file_path: str,
    sheet_name: str,
    start_row: int,
    name_col: int,
    metrics_col: int,
    metrics_codes: list,
    children_data: list,
    progress_callback=callable(lambda label, current_index, total_children: None),
):
    """
    children_data: { 'name': str, 'metric_code': score (1, 2, or 3) }
    """
    progress_callback("Loading the workbook", 0, 0)
    workbook = load_workbook(filename=file_path, read_only=False)
    sheet = workbook[sheet_name]
    current_row = start_row
    # metrics_col = name_col + 1
    for child in children_data:
        sheet.cell(row=current_row, column=name_col, value=child["name"])
        for metric_index, metric_code in enumerate(metrics_codes):
            score = child.get(metric_code)
            base_col = metrics_col + metric_index * 3

            # Clear previous values
            for offset in range(3):
                sheet.cell(row=current_row, column=base_col + offset).value = None

            # Set new value
            if score == 1:
                sheet.cell(row=current_row, column=base_col, value=1)
            elif score == 2:
                sheet.cell(row=current_row, column=base_col + 1, value=1)
            elif score == 3:
                sheet.cell(row=current_row, column=base_col + 2, value=1)
        progress_callback(
            child["name"], current_row - start_row + 1, len(children_data)
        )
        current_row += 1
        sleep(0.01)  # Simulate processing time
    return workbook


def get_table_boundaries(sheet):
    start_row, start_col = find_table_origin(sheet)

    footer_row = find_footer_row_index(sheet)
    if not footer_row:
        return None

    student_start_row = find_student_list_start_row(sheet, start_row, start_col)
    student_col = find_student_name_col_index(sheet, header_row=student_start_row - 2)
    last_col = find_last_data_col_index(sheet, footer_row)

    return {
        "start_row": start_row,
        "start_col": start_col,
        "student_start_row": student_start_row,
        "student_col": student_col,
        "last_student_row": footer_row - 1,
        "end_col": last_col,
    }


def reset_cell_to_default(cell):
    # Default font
    cell.font = Font(name="Calibri", size=11, bold=False, italic=False, color="000000")

    # Remove bg
    cell.fill = PatternFill(fill_type=None)

    # Remove borders
    cell.border = Border(
        left=Side(style=None),
        right=Side(style=None),
        top=Side(style=None),
        bottom=Side(style=None),
    )

    # Set default alignment
    cell.alignment = Alignment(horizontal="general", vertical="bottom")

    # Change number format (General / Non-Text)
    cell.number_format = "General"


def get_skip_cells(sheet):
    # Identify all merged ranges in the sheet
    merged_ranges = sheet.merged_cells.ranges

    # We collect the coordinates of their inner cells for easy searching as a list
    # We mark all cells except the top-left as "not to be touched"
    skip_cells = set()
    for rng in merged_ranges:
        # rng.cells will contain a tuple of all cells (row, col) in this range
        cells_list = list(rng.cells)
        if cells_list:
            top_left = cells_list[0]  # The first main cell (it can be cleared)
            for c in cells_list[1:]:
                skip_cells.add((c[0], c[1]))

    return skip_cells


def apply_complex_monitoring_borders(
    sheet,
    start_row,
    start_col,
    student_start_row,
    last_student_row,
    end_col,
    student_col,
):
    thin = Side(style="thin")
    medium = Side(style="medium")

    # Draw a thin line across the entire range (from Header to Footer)
    for row in sheet.iter_rows(
        min_row=start_row,
        max_row=last_student_row + 2,
        min_col=start_col,
        max_col=end_col,
    ):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Vertical medium lines
    important_columns = [start_col, student_col]
    for col in range(student_col + 1, end_col + 1, 3):
        important_columns.append(col)
    important_columns.append(end_col + 1)

    for col in important_columns:
        for row in range(start_row, last_student_row + 3):
            if col <= end_col:
                cell = sheet.cell(row=row, column=col)
                curr = cell.border
                cell.border = Border(
                    left=medium, right=curr.right, top=curr.top, bottom=curr.bottom
                )
            else:
                cell = sheet.cell(row=row, column=end_col)
                curr = cell.border
                cell.border = Border(
                    left=curr.left, right=medium, top=curr.top, bottom=curr.bottom
                )

    # Horizontal medium lines
    horizontal_medium_rows = [
        {"row": start_row, "from_col": start_col, "end_col": end_col},
        {"row": start_row + 1, "from_col": student_col + 1, "end_col": end_col},
        {
            "row": student_start_row - 3,
            "from_col": student_col + 1,
            "end_col": end_col,
        },
        {"row": student_start_row, "from_col": start_col, "end_col": end_col},
        {"row": last_student_row + 1, "from_col": start_col, "end_col": end_col},
        {"row": last_student_row + 3, "from_col": start_col, "end_col": end_col},
    ]

    for d in horizontal_medium_rows:
        for col in range(d["from_col"], d["end_col"] + 1):
            cell = sheet.cell(row=d["row"], column=col)
            curr = cell.border
            cell.border = Border(
                top=medium, left=curr.left, right=curr.right, bottom=curr.bottom
            )


def apply_monitoring_typography(
    sheet,
    start_row,
    start_col,
    student_start_row,
    last_student_row,
    end_col,
    student_col,
):
    skip_cells = get_skip_cells(sheet)
    # Table titles
    for row in range(1, start_row - 1):
        for col in range(1, end_col + 1):
            if (row, col) in skip_cells:
                continue
            cell = sheet.cell(row=row, column=col)
            if cell.value:
                cell.font = Font(name="Times New Roman", size=12, bold=True)
            else:
                reset_cell_to_default(cell)

    # Table headers
    for row in range(start_row, start_row + 1):
        for col in range(1, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.value:
                cell.font = Font(name="Times New Roman", size=12, bold=True)

    # Secondary table headers
    font_header = Font(name="Times New Roman", size=9, bold=False)
    for row in range(student_start_row - 3, student_start_row - 1):
        for col in range(student_col + 1, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.value:
                cell.font = font_header

    # Number and header
    cell = sheet.cell(row=student_start_row - 1, column=student_col)
    if cell.value:
        cell.font = font_header
    cell = sheet.cell(row=student_start_row - 1, column=student_col - 1)
    if cell.value:
        cell.font = font_header

    # Student list and Data Rows
    font_data = Font(name="Times New Roman", size=12, bold=False)
    for row in range(student_start_row, last_student_row + 1):
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.font = font_data

    # Footer
    font_footer = Font(name="Times New Roman", size=11, bold=True)
    for row in range(last_student_row + 1, last_student_row + 3):
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.value:
                cell.font = font_footer

    # Secondary tables
    font_secondary_tables = Font(name="Times New Roman", size=11, bold=False)
    for row in range(last_student_row + 3, last_student_row + 30):
        for col in range(1, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.value:
                cell.font = font_secondary_tables
            elif (cell.row, cell.column) not in skip_cells:
                reset_cell_to_default(cell)


def apply_monitoring_number_rounding(
    sheet,
    start_row,
    start_col,
    student_start_row,
    last_student_row,
    end_col,
    student_col,
):
    for row in range(last_student_row + 1, last_student_row + 30):
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.value is not None:
                val_str = str(cell.value)
                # If the cell contains a formula (for example, =SUM...) or a number
                if val_str.startswith("=") or isinstance(cell.value, (int, float)):
                    cell.number_format = "0"


def apply_monitoring_formula_fixing(
    sheet,
    start_row,
    start_col,
    student_start_row,
    last_student_row,
    end_col,
    student_col,
):
    actual_student_count = detect_actual_student_count(
        sheet, student_start_row, student_col
    )

    if actual_student_count <= 0:
        return

    for row in range(last_student_row + 1, last_student_row + 30):
        for col in range(1, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            if not (cell.value and str(cell.value).startswith("=")):
                continue
            formula = str(cell.value)
            updated_formula = formula

            # Template 1: =R[-1]C*100/24 → =E39*100/24
            if "*100/" in updated_formula:
                updated_formula = re.sub(
                    r"(=[A-Z]+\d+\*100/)\d+$",
                    f"\\g<1>{actual_student_count}",
                    updated_formula,
                )
            # Template 2: =R[-1]C/24% → =E39/24%
            elif "%" in updated_formula and "/" in updated_formula:
                updated_formula = re.sub(
                    r"(=[A-Z]+\d+/)\d+(%)",
                    f"\\g<1>{actual_student_count}\\g<2>",
                    updated_formula,
                )
            # Template 3: =RC[1]/100*24 → =F39/100*24
            elif "/100*" in updated_formula:
                updated_formula = re.sub(
                    r"(=[A-Z]+\d+/100\*)\d+$",
                    f"\\g<1>{actual_student_count}",
                    updated_formula,
                )

            if updated_formula != formula:
                cell.value = updated_formula


def remove_empty_rows_and_cols(
    sheet,
    start_row,
    start_col,
    student_start_row,
    last_student_row,
    end_col,
    student_col,
):
    # БОС ЖОЛДАРДЫ ТАЗАРТУ (Төменнен жоғары қарай жүреміз)
    current_row = last_student_row + 2

    while current_row >= start_row:
        row_is_empty = True

        # Осы жолдың ішінде бірде-бір мән бар-жоғын тексереміз
        for col in range(start_col, sheet.max_column + 1):
            if sheet.cell(row=current_row, column=col).value is not None:
                row_is_empty = False
                break

        # Егер жол таза бос болса - біздің қауіпсіз delete_row арқылы жоямыз
        if row_is_empty:
            delete_row(sheet, row_idx=current_row)

        current_row -= 1

    # БОС БАҒАНДАРДЫ ТАЗАРТУ (Оңнан солға қарай жүреміз)
    current_col = sheet.max_column
    while current_col >= start_col:
        col_is_empty = True

        # Осы бағанның ішінде мән бар-жоғын тексереміз
        for row in range(start_row, sheet.max_row + 1):
            if sheet.cell(row=row, column=current_col).value is not None:
                col_is_empty = False
                break

        # Егер баған бос болса - оны да жоямыз (openpyxl-де бағандарға арналған тура сондай әдіс бар)
        if col_is_empty:
            sheet.delete_cols(current_col, amount=1)

        current_col -= 1
