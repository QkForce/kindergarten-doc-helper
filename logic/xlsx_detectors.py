from openpyxl.worksheet.worksheet import Worksheet
from logic.loaders.utils import normalize_text


def find_table_origin(sheet, marker="№"):
    ref_row, ref_col = None, None
    for r in range(1, 31):
        for c in range(1, 11):
            val = sheet.cell(row=r, column=c).value
            if val and str(val).strip() == marker:
                ref_row, ref_col = r, c
                break
        if ref_row:
            break

    if not ref_row:
        return 4, 1
    start_row = ref_row
    for r in range(ref_row - 1, 1, -1):
        if sheet.cell(row=r, column=ref_col).border.top.style is not None:
            start_row = r
            return start_row, ref_col

    return start_row, ref_col


def find_student_list_start_row(sheet: Worksheet, start_row, start_col):
    for r in range(start_row, start_row + 20):
        val = sheet.cell(row=r, column=start_col).value
        if val is not None and str(val).strip() == "1":
            return r
    return start_row + 10


def find_student_name_col_index(sheet: Worksheet, header_row=12, marker="баланың аты"):
    for col in range(1, 10):
        val = sheet.cell(row=header_row, column=col).value
        if val and marker in str(val).lower():
            return col
    return 2


def detect_actual_student_count(
    sheet: Worksheet, student_start_row: int, student_col: int
) -> int:
    stop_words = ["барлығы", "қорытынды", "ескерту", "жоғары", "орташа", "төмен"]
    student_count = 0

    for row in range(student_start_row, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=student_col).value
        txt = normalize_text(cell_value)

        if not txt:
            continue
        if any(word in txt for word in stop_words):
            break

        student_count += 1

    return student_count


def find_footer_row_index(sheet: Worksheet, marker="барлығы"):
    for row in range(1, 200):
        val = sheet.cell(row=row, column=1).value
        if val and marker in str(val).lower():
            return row
    return None


def find_last_data_col_index(sheet: Worksheet, reference_row):
    curr_col = 2
    while True:
        if sheet.cell(row=reference_row, column=curr_col + 1).value is None:
            break
        curr_col += 1
    return curr_col
